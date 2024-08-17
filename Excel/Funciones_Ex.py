import openpyxl

class FunExcel():
    def __init__(self, driver):
        self.driver = driver
        
    def getRowCount(file, path, sheetName):
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetName]
        return (sheet.max_row)
    
    def getColumnCount(file, sheetName):
        Workbook = openpyxl.load_workbook(file)
        sheet = Workbook[sheetName]
        return (sheet.max_column)
    
    def readData(file, path, sheetName, rownum, columnno):
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetName]
        return sheet.cell(row=rownum, column=columnno).value
    
    def writeData(file, path, sheetName, rownum, columnno, data):
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetName]
        sheet.cell(row=rownum, column=columnno).value = data
        Workbook.save(path)
        print("Se guardo correctamente")